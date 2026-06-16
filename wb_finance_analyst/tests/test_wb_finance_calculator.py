import pandas as pd
from openpyxl import load_workbook

from wb_finance_analyst.config.settings import AppSettings
from wb_finance_analyst.domain.models import ColumnMap, CostItem, LoadedReport, ReportPeriod
from wb_finance_analyst.services.excel_exporter import ExcelExporter
from wb_finance_analyst.services.wb_finance_calculator import WBFinanceCalculator


class FakeCosts:
    def load(self):
        return {"A".casefold(): CostItem(product="A", cost=300)}


def test_wb_finance_calculator_reconciles_sales_returns_and_expenses(tmp_path):
    df = pd.DataFrame(
        [
            {"Тип документа": "Продажа", "Обоснование для оплаты": "Продажа", "Название": "A", "Кол-во": 1, "К перечислению Продавцу за реализованный Товар": 1000, "Услуги по доставке товара покупателю": 50, "Хранение": 2, "Удержания": 0, "Общая сумма штрафов": 0, "Операции на приемке": 0},
            {"Тип документа": "Возврат", "Обоснование для оплаты": "Возврат", "Название": "A", "Кол-во": 1, "К перечислению Продавцу за реализованный Товар": -200, "Услуги по доставке товара покупателю": 10, "Хранение": 1, "Удержания": 5, "Общая сумма штрафов": 3, "Операции на приемке": 4},
        ]
    )
    mapping = ColumnMap(
        doc_type="Тип документа",
        payment_reason="Обоснование для оплаты",
        product_name="Название",
        quantity="Кол-во",
        seller_transfer="К перечислению Продавцу за реализованный Товар",
        logistics="Услуги по доставке товара покупателю",
        storage="Хранение",
        deductions="Удержания",
        penalties="Общая сумма штрафов",
        acceptance="Операции на приемке",
    )
    loaded = LoadedReport(path=tmp_path / "raw.xlsx", dataframe=df, column_map=mapping, period=ReportPeriod())
    result = WBFinanceCalculator().calculate_loaded([loaded], AppSettings())
    assert result.summary["К перечислению по продажам"] == 1000
    assert result.summary["Возвраты"] == 200
    assert result.summary["К перечислению за товар"] == 800
    assert result.summary["Логистика"] == 60
    assert result.summary["Хранение"] == 3
    assert result.summary["Удержания/выплаты"] == 5
    assert result.summary["Штрафы"] == 3
    assert result.summary["Итого к оплате WB"] == 725


def test_return_compensation_is_goods_adjustment_not_sale_and_return(tmp_path):
    df = pd.DataFrame(
        [
            {"Тип документа": "Продажа", "Обоснование для оплаты": "Продажа", "Название": "A", "Кол-во": 1, "К перечислению Продавцу за реализованный Товар": 1000, "Услуги по доставке товара покупателю": 0, "Хранение": 0, "Удержания": 0, "Общая сумма штрафов": 0, "Операции на приемке": 0},
            {"Тип документа": "Возврат", "Обоснование для оплаты": "Возврат", "Название": "A", "Кол-во": 1, "К перечислению Продавцу за реализованный Товар": 100, "Услуги по доставке товара покупателю": 0, "Хранение": 0, "Удержания": 0, "Общая сумма штрафов": 0, "Операции на приемке": 0},
            {"Тип документа": "Возврат", "Обоснование для оплаты": "Добровольная компенсация при возврате", "Название": "A", "Кол-во": 0, "К перечислению Продавцу за реализованный Товар": 25, "Услуги по доставке товара покупателю": 0, "Хранение": 0, "Удержания": 0, "Общая сумма штрафов": 0, "Операции на приемке": 0},
        ]
    )
    mapping = ColumnMap(
        doc_type="Тип документа",
        payment_reason="Обоснование для оплаты",
        product_name="Название",
        quantity="Кол-во",
        seller_transfer="К перечислению Продавцу за реализованный Товар",
        logistics="Услуги по доставке товара покупателю",
        storage="Хранение",
        deductions="Удержания",
        penalties="Общая сумма штрафов",
        acceptance="Операции на приемке",
    )
    loaded = LoadedReport(path=tmp_path / "raw.xlsx", dataframe=df, column_map=mapping, period=ReportPeriod())
    result = WBFinanceCalculator().calculate_loaded([loaded], AppSettings())
    assert result.summary["К перечислению по продажам"] == 1000
    assert result.summary["Возвраты"] == 100
    assert result.summary["К перечислению за товар"] == 925
    assert result.summary["Итого к оплате WB"] == 925


def test_user_tables_are_grouped_and_excel_uses_readable_summary(tmp_path):
    df = pd.DataFrame(
        [
            {"Тип документа": "Продажа", "Обоснование для оплаты": "Продажа", "Название": "A", "Цена розничная": 1000, "К перечислению": 700},
            {"Тип документа": "Продажа", "Обоснование для оплаты": "Продажа", "Название": "A", "Цена розничная": 1000, "К перечислению": 650},
            {"Тип документа": "Возврат", "Обоснование для оплаты": "Возврат", "Название": "A", "Цена розничная": 1000, "К перечислению": 700},
        ]
    )
    mapping = ColumnMap(
        doc_type="Тип документа",
        payment_reason="Обоснование для оплаты",
        product_name="Название",
        retail_price="Цена розничная",
        wb_sold="Цена розничная",
        seller_transfer="К перечислению",
    )
    loaded = LoadedReport(path=tmp_path / "raw.xlsx", dataframe=df, column_map=mapping, period=ReportPeriod())
    result = WBFinanceCalculator(costs=FakeCosts()).calculate_loaded([loaded], AppSettings())

    assert list(result.reconciliation.columns) == ["Показатель", "Сумма"]
    assert list(result.management_profit.columns) == ["Показатель", "Сумма"]
    assert list(result.legend.columns) == ["Раздел", "Значение"]
    assert "report_kind" not in result.products.columns
    assert "Тип отчёта" not in result.products.columns
    assert "product" not in result.products.columns
    assert "nm_id" not in result.products.columns
    assert "Товар" in result.product_profit.columns
    assert result.product_profit.iloc[-1]["Товар"] == "Итого"
    assert list(result.sales.columns) == [
        "Товар",
        "Себестоимость/шт",
        "Розничная WB",
        "Количество",
        "WB перечисление",
        "Итого",
        "Средняя на шт",
        "Маржа",
        "Маржа %",
    ]
    assert len(result.sales) == 2
    sale_row = result.sales.iloc[0]
    total_row = result.sales.iloc[-1]
    assert sale_row["Товар"] == "A"
    assert sale_row["Количество"] == 2
    assert sale_row["WB перечисление"] == 1350
    assert sale_row["Маржа"] == 750
    assert total_row["Товар"] == "Итого"
    assert total_row["Количество"] == 2
    assert result.returns.iloc[0]["Количество"] == 1

    path = tmp_path / "report.xlsx"
    ExcelExporter().export(result, path)
    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames[:12] == [
        "Обозначения",
        "Общая сводка",
        "Управленческая прибыль",
        "Продажи",
        "Возвраты",
        "Товары",
        "Прибыль по товарам",
        "Реклама WB",
        "Внешние расходы",
        "Налоги",
        "Предупреждения",
        "Настройки отчёта",
    ]
    assert "Общая сводка" in wb.sheetnames
    assert "Сверка WB" not in wb.sheetnames
    assert "Графики" not in wb.sheetnames
    assert "Расходы WB" not in wb.sheetnames
    assert list(next(wb["Обозначения"].iter_rows(max_row=1, values_only=True))) == ["Раздел", "Значение"]
    assert list(next(wb["Общая сводка"].iter_rows(max_row=1, values_only=True))) == ["Показатель", "Сумма"]
    assert list(next(wb["Продажи"].iter_rows(max_row=1, values_only=True))) == list(result.sales.columns)
    assert "report_kind" not in list(next(wb["Товары"].iter_rows(max_row=1, values_only=True)))
    assert "nm_id" not in list(next(wb["Прибыль по товарам"].iter_rows(max_row=1, values_only=True)))
