from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from wb_finance_analyst.domain.constants import APP_NAME, APP_VERSION, TECH_SHEETS
from wb_finance_analyst.domain.models import WBFinanceResult


class ExcelExporter:
    def export(self, result: WBFinanceResult, path: Path) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            result.legend.to_excel(writer, sheet_name="Обозначения", index=False)
            result.reconciliation.to_excel(writer, sheet_name="Общая сводка", index=False)
            result.management_profit.to_excel(writer, sheet_name="Управленческая прибыль", index=False)
            result.sales.to_excel(writer, sheet_name="Продажи", index=False)
            result.returns.to_excel(writer, sheet_name="Возвраты", index=False)
            result.products.to_excel(writer, sheet_name="Товары", index=False)
            result.product_profit.to_excel(writer, sheet_name="Прибыль по товарам", index=False)
            result.ads.to_excel(writer, sheet_name="Реклама WB", index=False)
            result.expenses.to_excel(writer, sheet_name="Внешние расходы", index=False)
            result.taxes.to_excel(writer, sheet_name="Налоги", index=False)
            pd.DataFrame({"Предупреждение": result.warnings}).to_excel(writer, sheet_name="Предупреждения", index=False)
            result.report_settings.to_excel(writer, sheet_name="Настройки отчёта", index=False)
            result.reconciliation.to_excel(writer, sheet_name="_DATA_SUMMARY", index=False)
            result.product_profit.to_excel(writer, sheet_name="_DATA_PRODUCTS", index=False)
            result.ads.to_excel(writer, sheet_name="_DATA_ADS", index=False)
            result.operations.to_excel(writer, sheet_name="_DATA_OPERATIONS", index=False)
            result.expenses.to_excel(writer, sheet_name="_DATA_EXPENSES", index=False)
            self._meta(result).to_excel(writer, sheet_name="_DATA_META", index=False)

        wb = load_workbook(path)
        for ws in wb.worksheets:
            self._format_sheet(ws)
            if ws.title in TECH_SHEETS:
                ws.sheet_state = "hidden"
        wb.save(path)
        return path

    def default_path(self, directory: Path, result: WBFinanceResult) -> Path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        period = result.period.label.replace(" ", "").replace("-", "_").replace(".", "")
        return directory / f"WB_Finance_{period}_{stamp}.xlsx"

    def _meta(self, result: WBFinanceResult) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {"key": "app_name", "value": APP_NAME},
                {"key": "app_version", "value": APP_VERSION},
                {"key": "report_id", "value": result.report_id},
                {"key": "period_start", "value": result.period.start.isoformat() if result.period.start else ""},
                {"key": "period_end", "value": result.period.end.isoformat() if result.period.end else ""},
                {"key": "generated_at", "value": result.generated_at.isoformat()},
                {"key": "source_files", "value": "; ".join(result.source_files)},
                {"key": "settings_hash", "value": result.settings_hash},
                {"key": "data_source", "value": "WB API" if "WB API" in result.source_files else "Excel"},
            ]
        )

    def _wb_expenses(self, result: WBFinanceResult) -> pd.DataFrame:
        names = ["Логистика", "Хранение", "Удержания/выплаты", "Штрафы", "Операции при приемке", "Лояльность"]
        return pd.DataFrame([{"Расход": name, "Сумма": result.summary.get(name, 0.0)} for name in names])

    def _format_sheet(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="EAF1F8")
        total_fill = PatternFill("solid", fgColor="F7F9D7")
        border = Border(bottom=Side(style="thin", color="D9E2EC"))
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1F2937")
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    if str(ws.cell(row=1, column=cell.column).value).endswith("%"):
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = '# ##0.00'
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = Font(color="B91C1C")
            first = " ".join(str(cell.value or "") for cell in row[:2])
            if first.startswith("Итого") or first.startswith("Чистая"):
                for cell in row:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
        for column_cells in ws.columns:
            max_len = 8
            for cell in column_cells:
                text = str(cell.value or "")
                max_len = max(max_len, min(len(text), 60))
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_len + 2

    def _add_charts_sheet(self, wb) -> None:
        if "Графики" in wb.sheetnames:
            del wb["Графики"]
        ws = wb.create_sheet("Графики")
        source = wb["Общая сводка"]
        ws["A1"] = "Основные показатели"
        ws["A1"].font = Font(bold=True, size=14)
        chart = BarChart()
        chart.title = "Общая сводка"
        chart.y_axis.title = "Руб."
        chart.x_axis.title = "Показатель"
        max_row = min(source.max_row, 10)
        data = Reference(source, min_col=2, min_row=1, max_row=max_row)
        cats = Reference(source, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 24
        ws.add_chart(chart, "A3")
